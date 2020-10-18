''' 
author:紫夏
Time:2020/8/30 17:07
'''

'''封装的需求
1、封装一个读数据的功能
    读数据需要用到什么？ excel文件路径，表单名


封装的数据


'''
import openpyxl

class Handle_excel:
    def __init__(self,filename,sheetname):
        self.filename=filename
        self.sheetname=sheetname

    def read_data(self):
        wb=openpyxl.load_workbook(self.filename)
        sh=wb[self.sheetname]

        cases_data=[]
        rows_data=list(sh.rows)
        title=[]
        for i in rows_data[0]:
            title.append(i.value)

        for i in rows_data[1:]:
            # 这一行放在循环里面和循环外面差别这么大。。放在循环外面会都是同一个value
            values = []
            for j in i:
                values.append(j.value)
            case=dict(zip(title,values))
            cases_data.append(case)
        return cases_data

    def write_data(self,row,column,value):
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname]
        sh.cell(row=row,column=column,value=value)
        wb.save(self.filename)


if __name__ == '__main__':
    # excel=Handle_excel('cases.xlsx','register')
    excel=Handle_excel('cases.xlsx','login')
    # cases=excel.read_data()
    # print(cases)
    excel.write_data(1,1,'测试1')
    excel.write_data(2,1,'测试2')
    excel.write_data(1,2,'测试3')


